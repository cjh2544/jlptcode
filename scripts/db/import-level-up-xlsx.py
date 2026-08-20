"""
Replace local MariaDB level_up rows from the Excel export.
Does not touch MongoDB.
"""
from __future__ import annotations

import os
import secrets
from pathlib import Path
from urllib.parse import unquote, urlparse

import pymysql
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]


def load_env():
    for name in (".env.local", ".env"):
        path = ROOT / name
        if not path.exists():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


load_env()

EXCEL_PATH = Path(r"C:\Users\xodus\Downloads\(260807)moku.level_up.xlsx")
CHUNK = 200


def as_str(value, fallback=None):
    if value is None or value == "":
        return fallback
    return str(value).strip()


def as_int(value, fallback=None):
    if value is None or value == "":
        return fallback
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return fallback


def as_text(value):
    if value is None or value == "":
        return None
    return str(value)


def parse_db_url(url: str):
    parsed = urlparse(url)
    return {
        "host": parsed.hostname or "127.0.0.1",
        "port": parsed.port or 3306,
        "user": unquote(parsed.username or ""),
        "password": unquote(parsed.password or ""),
        "database": parsed.path.lstrip("/"),
        "charset": "utf8mb4",
    }


def build_question(row):
    content = row.get("question.content")
    content_org = row.get("question.contentOrg")
    audio_link = as_str(row.get("question.audio.link"))
    image_link = as_str(row.get("question.image.link"))
    question = {}
    if content not in (None, ""):
        question["content"] = str(content)
    if content_org not in (None, ""):
        question["contentOrg"] = str(content_org)
    if audio_link:
        question["audio"] = {"link": audio_link}
    if image_link:
        question["image"] = {"link": image_link}
    return question or {}


def build_choices(row):
    choices = []
    for idx in range(4):
        no = as_int(row.get(f"choices[{idx}].no"))
        content = row.get(f"choices[{idx}].content")
        if no is None and (content is None or content == ""):
            continue
        choices.append({"no": no, "content": None if content is None else str(content)})
    return choices or None


def build_sentence(row):
    translation = row.get("sentence.translation")
    reading = row.get("sentence.reading")
    if translation in (None, "") and reading in (None, ""):
        return None
    sentence = {}
    if translation not in (None, ""):
        sentence["translation"] = str(translation)
    if reading not in (None, ""):
        sentence["reading"] = str(reading)
    return sentence


def build_locale(row):
    locale = {}
    for key in ("en", "cn", "my"):
        value = row.get(f"sentence_locale.{key}")
        if value not in (None, ""):
            locale[key] = str(value)
    return locale or None


def row_to_parent(headers, values, fallback_id):
    row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
    raw_id = row.get("_id")
    record_id = str(int(raw_id)) if isinstance(raw_id, (int, float)) else as_str(raw_id, str(fallback_id))
    question = build_question(row)
    sentence = build_sentence(row) or {}
    locale = build_locale(row) or {}
    audio = question.get("audio") or {}
    image = question.get("image") or {}
    parent = (
        record_id[:24],
        as_str(row.get("year"), ""),
        as_str(row.get("level"), ""),
        as_int(row.get("sortNo"), 0),
        as_str(row.get("classification"), ""),
        as_int(row.get("questionNo")),
        as_int(row.get("questionGroupNo")),
        as_int(row.get("questionContentNo")),
        as_text(question.get("content")),
        as_text(question.get("contentOrg")),
        as_text(audio.get("link")),
        as_text(image.get("link")),
        as_str(row.get("questionGroupType"), ""),
        as_str(row.get("questionType")),
        as_int(row.get("answer")),
        as_text(sentence.get("translation")),
        as_text(sentence.get("reading")),
        as_str(row.get("speaker")),
        as_text(locale.get("en")),
        as_text(locale.get("cn")),
        as_text(locale.get("my")),
    )
    choices = [
        (secrets.token_hex(12), record_id[:24], item["no"], item["content"])
        for item in (build_choices(row) or [])
    ]
    return parent, choices


def main():
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        raise SystemExit("DATABASE_URL is required")
    if not EXCEL_PATH.exists():
        raise SystemExit(f"Excel not found: {EXCEL_PATH}")

    conn = pymysql.connect(**parse_db_url(db_url), autocommit=False)
    try:
        with conn.cursor() as cur:
            cur.execute("ALTER TABLE `level_up` MODIFY `speaker` TEXT NULL")
            cur.execute("SELECT COUNT(*) FROM `level_up`")
            before = cur.fetchone()[0]
            cur.execute("DELETE FROM `level_up`")
            print(f"deleted_level_up {before}")

            wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = None
            parents = []
            choices = []
            inserted = 0
            skipped = 0
            parent_sql = """
                INSERT INTO `level_up` (
                  `id`, `year`, `level`, `sortNo`, `classification`, `questionNo`,
                  `questionGroupNo`, `questionContentNo`, `question_content`, `question_content_org`,
                  `question_audio_link`, `question_image_link`, `questionGroupType`, `questionType`,
                  `answer`, `sentence_translation`, `sentence_reading`, `speaker`,
                  `sentence_locale_en`, `sentence_locale_cn`, `sentence_locale_my`,
                  `created_at`, `updated_at`
                ) VALUES (
                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(3), NOW(3)
                )
            """
            choice_sql = """
                INSERT INTO `level_up_choice` (`id`, `parent_id`, `no`, `content`)
                VALUES (%s, %s, %s, %s)
            """
            for i, values in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c is not None else "" for c in values]
                    continue
                if values[0] is None and values[1] is None:
                    skipped += 1
                    continue
                parent, row_choices = row_to_parent(headers, values, i)
                parents.append(parent)
                choices.extend(row_choices)
                if len(parents) >= CHUNK:
                    cur.executemany(parent_sql, parents)
                    if choices:
                        cur.executemany(choice_sql, choices)
                    inserted += len(parents)
                    parents = []
                    choices = []
                    print(f"inserted {inserted}")
            if parents:
                cur.executemany(parent_sql, parents)
                if choices:
                    cur.executemany(choice_sql, choices)
                inserted += len(parents)
            wb.close()
            conn.commit()
            cur.execute("SELECT COUNT(*) FROM `level_up`")
            after = cur.fetchone()[0]
            print(f"done inserted={inserted} skipped={skipped} table_count={after}")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
